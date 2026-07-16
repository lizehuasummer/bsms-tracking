"""
BSMS GitLab → Excel 每周同步脚本

用法:
  python sync_gitlab_to_excel.py              # 使用缓存（推荐）
  python sync_gitlab_to_excel.py --fresh       # 删除缓存，全部重新抓取
  python sync_gitlab_to_excel.py --excel-only  # 跳过 GitLab 抓取，仅用缓存填表

执行流程:
  1. 抓取 GitLab Function::dev + Bug::dev 的所有 issue
  2. 对每个 issue 抓取 system notes（状态变更）和 label events（Process::pass）
  3. 匹配 Excel 三个 Sheet 并填写
  4. 保存 Excel

详细说明见 README.md
"""
import json
import re
import time
import os
import sys
import requests
import urllib3
from concurrent.futures import ThreadPoolExecutor, as_completed
urllib3.disable_warnings()

# ============ 配置 ============
GITLAB_URL = "https://gitlab.stpass.com"
GITLAB_TOKEN = "glpat-3HdBPr8u1xJaNy23qLzWzm86MQp1OnEH.01.0w0o8vjhx"
PROJECT_ID = 4
EXCEL_PATH = r"D:\Users\LZH\Documents\opencode\bsms_tracking\【交付HK】BSMS项目整体进度计划 (更新中)20260702备份_new.xlsx"
OUTPUT_PATH = r"D:\Users\LZH\Documents\opencode\bsms_tracking\【交付HK】BSMS项目整体进度计划 (更新中)20260702备份_new.xlsx"
# 如果 OUTPUT_PATH 被占用（Excel 打开中），脚本会自动加 _new 后缀
CACHE_DIR = r"C:\Users\EDY\AppData\Local\Temp\opencode"
ISSUES_CACHE = os.path.join(CACHE_DIR, "gitlab_issues_cache.json")
NOTES_CACHE = os.path.join(CACHE_DIR, "gitlab_notes_cache.json")
CONCURRENT_WORKERS = 10

session = requests.Session()
session.headers.update({"Private-Token": GITLAB_TOKEN})
session.verify = False


# ============ GitLab API ============
def fetch_all_issues(label, state="all"):
    all_issues = []
    page = 1
    while True:
        r = session.get(
            f"{GITLAB_URL}/api/v4/projects/{PROJECT_ID}/issues",
            params={"labels": label, "state": state, "per_page": 100, "page": page},
            timeout=30,
        )
        r.raise_for_status()
        data = r.json()
        if not data:
            break
        all_issues.extend(data)
        total_pages = int(r.headers.get("X-Total-Pages", "1"))
        print(f"  {label} page {page}/{total_pages} ({len(all_issues)}/{r.headers.get('X-Total', '?')})")
        if page >= total_pages:
            break
        page += 1
    return all_issues


def fetch_issue_notes(iid):
    key = str(iid)
    try:
        r1 = session.get(
            f"{GITLAB_URL}/api/v4/projects/{PROJECT_ID}/issues/{iid}/notes",
            params={"per_page": 100, "sort": "asc", "order_by": "created_at"},
            timeout=15,
        )
        r1.raise_for_status()
        r2 = session.get(
            f"{GITLAB_URL}/api/v4/projects/{PROJECT_ID}/issues/{iid}/resource_label_events",
            params={"per_page": 100},
            timeout=15,
        )
        r2.raise_for_status()
        return key, {"notes": r1.json(), "label_events": r2.json()}
    except Exception as e:
        print(f"  ERR iid {iid}: {e}")
        return key, {"notes": [], "label_events": []}


# ============ 解析工具 ============
def fmt_time(t):
    if not t:
        return None
    return t[:16].replace("T", " ")


def extract_feature_id(title):
    if not title:
        return None
    m = re.search(r"([A-Z]{2,5}-[A-Z]{1,5}\d{1,4}(?:-\d+)?)", title)
    return m.group(1) if m else None


def parse_status_from_notes(notes):
    changes = []
    for note in notes:
        if not note.get("system"):
            continue
        m = re.search(r"set status to \*\*(.+?)\*\*", note.get("body", ""))
        if m:
            changes.append({
                "status": m.group(1),
                "time": note["created_at"],
                "user": note.get("author", {}).get("username", ""),
            })
    return changes


def parse_processpass_from_events(events):
    result = []
    for ev in events:
        if ev.get("label", {}).get("name") == "Process::pass" and ev.get("action") == "add":
            result.append({"time": ev["created_at"], "user": ev.get("user", {}).get("username", "")})
    return result


def find_last(items, status_name):
    result = None
    for item in items:
        if item["status"] == status_name:
            result = item
    return result


# ============ 主流程 ============
def step1_fetch_issues():
    cache_file = ISSUES_CACHE
    if os.path.exists(cache_file):
        mtime = os.path.getmtime(cache_file)
        if time.time() - mtime < 3600:
            print("[1] 加载 issues 缓存...")
            with open(cache_file, "r", encoding="utf-8") as f:
                cache = json.load(f)
            return cache.get("functiondev", []), cache.get("bugdev", [])

    print("[1] 抓取 GitLab issues...")
    func_issues = fetch_all_issues("Function::dev")
    print(f"  Function::dev: {len(func_issues)}")
    bug_issues = fetch_all_issues("Bug::dev")
    print(f"  Bug::dev: {len(bug_issues)}")

    with open(cache_file, "w", encoding="utf-8") as f:
        json.dump({"functiondev": func_issues, "bugdev": bug_issues}, f, ensure_ascii=False)

    return func_issues, bug_issues


def step2_fetch_notes(func_issues, bug_issues):
    print("[2] 抓取 issue notes + label events...")

    if os.path.exists(NOTES_CACHE):
        with open(NOTES_CACHE, "r", encoding="utf-8") as f:
            notes_cache = json.load(f)
    else:
        notes_cache = {}

    all_iids = [issue["iid"] for issue in func_issues + bug_issues]
    to_fetch = [iid for iid in all_iids if str(iid) not in notes_cache]
    print(f"  已缓存: {len(notes_cache)}, 需抓取: {len(to_fetch)}")

    if not to_fetch:
        print("  全部已缓存，跳过")
        return notes_cache

    fetched = 0
    with ThreadPoolExecutor(max_workers=CONCURRENT_WORKERS) as executor:
        futures = {executor.submit(fetch_issue_notes, iid): iid for iid in to_fetch}
        for future in as_completed(futures):
            key, data = future.result()
            notes_cache[key] = data
            fetched += 1
            if fetched % 100 == 0:
                print(f"  进度: {fetched}/{len(to_fetch)}")
                with open(NOTES_CACHE, "w", encoding="utf-8") as f:
                    json.dump(notes_cache, f, ensure_ascii=False)

    with open(NOTES_CACHE, "w", encoding="utf-8") as f:
        json.dump(notes_cache, f, ensure_ascii=False)
    print(f"  完成，总缓存: {len(notes_cache)}")

    return notes_cache


def step3_fill_excel(func_issues, bug_issues, notes_cache):
    print("[3] 填写 Excel...")
    import openpyxl

    func_by_fid = {}
    for issue in func_issues:
        fid = extract_feature_id(issue.get("title", ""))
        if fid:
            func_by_fid.setdefault(fid, []).append(issue)

    print(f"  Function::dev 功能点编号: {len(func_by_fid)} 个唯一值")

    wb = openpyxl.load_workbook(EXCEL_PATH)

    def find_sheet(keyword):
        for name in wb.sheetnames:
            if keyword.lower() in name.lower():
                return wb[name]
        raise ValueError(f"找不到包含 '{keyword}' 的 sheet")

    # --- Sheet 1: Sprint详细安排 ---
    ws1 = find_sheet("Sprint")
    print(f"\n  Sheet 1: {ws1.title} ({ws1.max_row} rows)")
    matched1 = 0
    for row in range(5, ws1.max_row + 1):
        fid = ws1.cell(row, 5).value
        fname = ws1.cell(row, 6).value
        if not fid:
            continue
        fid = str(fid).strip()
        candidates = func_by_fid.get(fid, [])
        if not candidates:
            continue

        matched_issue = candidates[0]
        if len(candidates) > 1 and fname:
            for c in candidates:
                if str(fname).strip() in c.get("title", ""):
                    matched_issue = c
                    break

        key = str(matched_issue["iid"])
        cdata = notes_cache.get(key, {"notes": [], "label_events": []})
        status_changes = parse_status_from_notes(cdata.get("notes", []))
        pass_events = parse_processpass_from_events(cdata.get("label_events", []))
        current_labels = set(matched_issue.get("labels", []))

        # I: 最新状态
        latest_status = status_changes[-1]["status"] if status_changes else None
        if latest_status == "测试中" and "Process::pass" in current_labels:
            latest_status = "processpass"
        if latest_status:
            ws1.cell(row, 9).value = latest_status

        # L/M: 测试中
        test_sc = find_last(status_changes, "测试中")
        if test_sc:
            ws1.cell(row, 13).value = fmt_time(test_sc["time"])
            ws1.cell(row, 12).value = test_sc["user"]

        # N/O: Process::pass
        if pass_events:
            ws1.cell(row, 15).value = fmt_time(pass_events[-1]["time"])
            ws1.cell(row, 14).value = pass_events[-1]["user"]

        # R/S: 已完成
        done_sc = find_last(status_changes, "已完成")
        if done_sc:
            ws1.cell(row, 19).value = fmt_time(done_sc["time"])
            ws1.cell(row, 18).value = done_sc["user"]

        matched1 += 1
    print(f"  匹配: {matched1}")

    # --- Sheet 2: EX开发功能点 ---
    ws2 = find_sheet("EX")
    print(f"\n  Sheet 2: {ws2.title} ({ws2.max_row} rows)")
    matched2 = 0
    for row in range(5, ws2.max_row + 1):
        fid = ws2.cell(row, 5).value
        fname = ws2.cell(row, 6).value
        if not fid:
            continue
        fid = str(fid).strip()
        candidates = func_by_fid.get(fid, [])
        if not candidates:
            continue

        matched_issue = candidates[0]
        if len(candidates) > 1 and fname:
            for c in candidates:
                if str(fname).strip() in c.get("title", ""):
                    matched_issue = c
                    break

        key = str(matched_issue["iid"])
        cdata = notes_cache.get(key, {"notes": [], "label_events": []})
        status_changes = parse_status_from_notes(cdata.get("notes", []))
        pass_events = parse_processpass_from_events(cdata.get("label_events", []))

        test_sc = find_last(status_changes, "测试中")
        if test_sc:
            ws2.cell(row, 13).value = fmt_time(test_sc["time"])
            ws2.cell(row, 12).value = test_sc["user"]

        if pass_events:
            ws2.cell(row, 15).value = fmt_time(pass_events[-1]["time"])
            ws2.cell(row, 14).value = pass_events[-1]["user"]

        done_sc = find_last(status_changes, "已完成")
        if done_sc:
            ws2.cell(row, 19).value = fmt_time(done_sc["time"])
            ws2.cell(row, 18).value = done_sc["user"]

        matched2 += 1
    print(f"  匹配: {matched2}")

    # --- Sheet 3: bug ---
    ws3 = find_sheet("bug")
    print(f"\n  Sheet 3: {ws3.title}")
    for row in range(2, ws3.max_row + 1):
        for col in range(1, 14):
            cell = ws3.cell(row, col)
            if not isinstance(cell, openpyxl.cell.cell.MergedCell):
                cell.value = None

    bug_row = 2
    for issue in bug_issues:
        key = str(issue["iid"])
        cdata = notes_cache.get(key, {"notes": [], "label_events": []})
        status_changes = parse_status_from_notes(cdata.get("notes", []))

        subsystem = priority = None
        for label in issue.get("labels", []):
            if label.startswith("BSMS::"):
                subsystem = label.replace("BSMS::", "")
            elif label.startswith("Priority::"):
                priority = label.replace("Priority::", "")

        latest_status = status_changes[-1]["status"] if status_changes else issue.get("state", "")

        fix_sc = find_last(status_changes, "待修复")
        test_sc = find_last(status_changes, "测试中")
        done_sc = find_last(status_changes, "已完成")

        ws3.cell(bug_row, 1).value = subsystem
        ws3.cell(bug_row, 2).value = f"#{issue['iid']}"
        ws3.cell(bug_row, 3).value = issue.get("title", "")
        ws3.cell(bug_row, 4).value = priority
        ws3.cell(bug_row, 5).value = latest_status
        ws3.cell(bug_row, 6).value = issue.get("author", {}).get("username", "")
        ws3.cell(bug_row, 7).value = fmt_time(issue.get("created_at"))

        if fix_sc:
            ws3.cell(bug_row, 8).value = fix_sc["user"]
            ws3.cell(bug_row, 9).value = fmt_time(fix_sc["time"])
        if test_sc:
            ws3.cell(bug_row, 10).value = test_sc["user"]
            ws3.cell(bug_row, 11).value = fmt_time(test_sc["time"])
        if done_sc:
            ws3.cell(bug_row, 12).value = done_sc["user"]
            ws3.cell(bug_row, 13).value = fmt_time(done_sc["time"])

        bug_row += 1
    print(f"  Bug 行数: {bug_row - 2}")

    try:
        wb.save(OUTPUT_PATH)
        print(f"\n  已保存: {OUTPUT_PATH}")
    except PermissionError:
        alt_path = OUTPUT_PATH.replace(".xlsx", "_new.xlsx")
        wb.save(alt_path)
        print(f"\n  原文件被占用，已保存到: {alt_path}")


def main():
    fresh = "--fresh" in sys.argv
    excel_only = "--excel-only" in sys.argv

    if fresh:
        for f in [ISSUES_CACHE, NOTES_CACHE]:
            if os.path.exists(f):
                os.remove(f)
                print(f"已删除缓存: {f}")

    if not excel_only:
        func_issues, bug_issues = step1_fetch_issues()
        notes_cache = step2_fetch_notes(func_issues, bug_issues)
    else:
        print("[1] 仅使用缓存填表")
        with open(ISSUES_CACHE, "r", encoding="utf-8") as f:
            ic = json.load(f)
        func_issues = ic["functiondev"]
        bug_issues = ic["bugdev"]
        with open(NOTES_CACHE, "r", encoding="utf-8") as f:
            notes_cache = json.load(f)

    print(f"\n  Function::dev: {len(func_issues)}, Bug::dev: {len(bug_issues)}, Notes: {len(notes_cache)}")

    step3_fill_excel(func_issues, bug_issues, notes_cache)
    print("\n完成！")


if __name__ == "__main__":
    main()
