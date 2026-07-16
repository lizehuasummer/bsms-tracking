"""
在看板 Excel 末尾添加一个 sheet，记录「开发每日汇总看板」和「测试每日汇总看板」每一列的统计口径。
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

EXCEL_PATH = r"D:\Users\LZH\Documents\opencode\bsms_tracking\【交付HK】BSMS项目整体进度计划 (更新中)20260702备份_new.xlsx"

wb = openpyxl.load_workbook(EXCEL_PATH)

sheet_name = "看板列说明"
if sheet_name in wb.sheetnames:
    del wb[sheet_name]
ws = wb.create_sheet(sheet_name)

# ===== 样式 =====
title_font = Font(name="微软雅黑", size=14, bold=True, color="FFFFFF")
section_font = Font(name="微软雅黑", size=12, bold=True, color="FFFFFF")
header_font = Font(name="微软雅黑", size=10, bold=True)
body_font = Font(name="微软雅黑", size=10)
title_fill = PatternFill("solid", fgColor="2F5496")
section_fill = PatternFill("solid", fgColor="4472C4")
header_fill = PatternFill("solid", fgColor="D6E4F0")
dev_fill = PatternFill("solid", fgColor="E2EFDA")
test_fill = PatternFill("solid", fgColor="FCE4D6")
wrap = Alignment(wrap_text=True, vertical="top")
center = Alignment(horizontal="center", vertical="center", wrap_text=True)
thin_border = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)

# ===== 通用说明 =====
row = 1
ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
cell = ws.cell(row, 1, "BSMS 每日汇总看板 — 列统计口径说明")
cell.font = title_font
cell.fill = title_fill
cell.alignment = center
ws.row_dimensions[row].height = 30
row += 1

ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
cell = ws.cell(row, 1, "数据来源：GitLab 项目 aladdinx/document/dev-design (ID=4) 的 Issue system notes + label events\n"
                        "日期范围：从最早的 issue 创建日期到今天，按天连续排列\n"
                        "前端用户组：池一帆、张永辉、张映杰、丁传琨、蔡骁聪、许崇楠、姚川、林河、刘浩龙、李术扬\n"
                        "后端用户组：邱靖凯、欧军豪、肖文浩、潘智豪、徐学凤、钟浩锋、蔡晓钰、陈丽霞、郑宝华、邓康、何奎")
cell.font = body_font
cell.alignment = wrap
ws.row_dimensions[row].height = 60
row += 2

# ===== 前后端分类规则 =====
ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
cell = ws.cell(row, 1, "BUG 前/后端分类规则（适用于开发看板 Q/R/S 列 和 测试看板 T/U/V 列）")
cell.font = section_font
cell.fill = section_fill
cell.alignment = center
ws.row_dimensions[row].height = 24
row += 1

rules = [
    "1. 如果 issue 带 Port::fullstack 标签 → 归入「前+后端」",
    "2. 如果在本次操作之前，前端用户和后端用户都参与过该 issue 的活动 → 归入「前+后端」",
    "3. 如果操作人是前端用户：之前有后端用户活动 →「前+后端」；否则 →「仅前端」",
    "4. 如果操作人是后端用户：之前有前端用户活动 →「前+后端」；否则 →「仅后端」",
    "5. 如果 issue 带 Port::frontend 但无 Port::backend →「仅前端」",
    "6. 如果 issue 带 Port::backend 但无 Port::frontend →「仅后端」",
    "7. 以上都不满足 → 归入「前+后端」",
]
for rule in rules:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    cell = ws.cell(row, 1, rule)
    cell.font = body_font
    cell.alignment = wrap
    row += 1
row += 1

# ===== 表头 =====
headers = ["列号", "列名", "统计口径", "数据来源", "备注"]
for ci, h in enumerate(headers, 1):
    cell = ws.cell(row, ci, h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center
    cell.border = thin_border
ws.row_dimensions[row].height = 22
header_row = row
row += 1

# ===== 开发每日汇总看板 =====
ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
cell = ws.cell(row, 1, "一、开发每日汇总看板（19 列）")
cell.font = section_font
cell.fill = section_fill
cell.alignment = center
ws.row_dimensions[row].height = 24
row += 1

dev_rows = [
    ("A", "日期", "当天的日期（YYYY-MM-DD）", "系统日期", "从最早 issue 创建日到今天，按天连续"),
    ("B", "当日新增功能需求", "当天创建的 Function::dev issue 数量", "issue.created_at 的日期 = 当天", ""),
    ("C", "当日前端完成开发数", "当天添加 front::finished 标签、且操作人是前端用户的 Function::dev issue 数", "label_event: label=front::finished, action=add, user ∈ 前端用户组", "每个 issue 当天最多计 1 次"),
    ("D", "当日后端完成开发数", "当天添加 backend::finished 标签、且操作人是后端用户的 Function::dev issue 数", "label_event: label=backend::finished, action=add, user ∈ 后端用户组", "每个 issue 当天最多计 1 次"),
    ("E", "当日联调完成数", "当天状态变更为「测试中」的 Function::dev issue 数", "system note: body 含 'set status to **测试中**'", "表示开发提交测试"),
    ("F", "当日联调完成数", "同 E 列（重复列）", "同 E", "E=F"),
    ("G", "累计提交测试总数", "截至当天，处于「已关闭」状态 或 最新状态为「测试中」的 Function::dev issue 累计数", "遍历每个 issue 的时间线，取当天及之前的事件判断", "= 已关闭数 + 当前测试中数"),
    ("H", "累计提交测试完成率", "当天已完成的 issue 数 / G 列值 × 100%", "已完成 = 状态变更为「已完成」 或 issue 已关闭", "百分比格式"),
    ("I", "当日提交P0", "当天状态变更为「测试中」且带 Priority::P0 标签的 Bug::dev issue 数", "system note + issue labels", ""),
    ("J", "当日提交P1", "同上，Priority::P1", "同上", ""),
    ("K", "当日提交P2", "同上，Priority::P2", "同上", ""),
    ("L", "当日提交P3", "同上，Priority::P3", "同上", ""),
    ("M", "当日提交P4", "同上，Priority::P4", "同上", ""),
    ("N", "当日提交合计", "I+J+K+L+M", "计算列", ""),
    ("O", "当日剩余待修复Bug总量", "当天处于 open 状态（创建日 ≤ 当天 且 未关闭）且最新系统状态为「待修复」的 Bug::dev issue 数", "issue.created_at / closed_at + system note 状态", ""),
    ("P", "高级别(P0+P1)剩余存量", "当天 open 且带 Priority::P0 或 Priority::P1 的 Bug::dev issue 数", "同 O 列逻辑 + Priority 过滤", ""),
    ("Q", "仅前端BUG当日提交测试验证数", "当天状态变更为「测试中」的 Bug::dev issue 中，按分类规则判定为「仅前端」的数量", "system note + 前后端分类规则", "Q+R+S = N"),
    ("R", "仅后端BUG当日提交测试验证数", "同上，判定为「仅后端」", "同上", ""),
    ("S", "前+后端BUG当日提交测试验证数", "同上，判定为「前+后端」", "同上", ""),
]

for r_data in dev_rows:
    for ci, val in enumerate(r_data, 1):
        cell = ws.cell(row, ci, val)
        cell.font = body_font
        cell.alignment = wrap
        cell.border = thin_border
        cell.fill = dev_fill
    row += 1

row += 1

# ===== 测试每日汇总看板 =====
ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
cell = ws.cell(row, 1, "二、测试每日汇总看板（23 列）")
cell.font = section_font
cell.fill = section_fill
cell.alignment = center
ws.row_dimensions[row].height = 24
row += 1

test_rows = [
    ("A", "日期", "当天的日期（YYYY-MM-DD）", "系统日期", "从最早 issue 创建日到今天，按天连续"),
    ("B", "当日新提交功能测试", "当天状态变更为「测试中」的 Function::dev issue 数", "system note: 'set status to **测试中**'", "开发提交给测试"),
    ("C", "当日完成功能测试", "当天状态变更为「已完成」的 Function::dev issue 数", "system note: 'set status to **已完成**'", "测试通过完成"),
    ("D", "当日剩余功能测试", "截至当天，处于 open 状态且最新状态为「测试中」的 Function::dev issue 数", "遍历 issue 时间线取最新状态", "正在测试中的存量"),
    ("E", "当日新增P0", "当天添加 Priority::P0 标签的 Bug::dev issue 数", "label_event: label=Priority::P0, action=add", ""),
    ("F", "当日新增P1", "同上，Priority::P1", "同上", ""),
    ("G", "当日新增P2", "同上，Priority::P2", "同上", ""),
    ("H", "当日新增P3", "同上，Priority::P3", "同上", ""),
    ("I", "当日新增P4", "同上，Priority::P4", "同上", ""),
    ("J", "当日新增合计", "E+F+G+H+I", "计算列", ""),
    ("K", "当日完成P0", "当天状态变更为「已完成」且带 Priority::P0 标签的 Bug::dev issue 数", "system note + issue labels", ""),
    ("L", "当日完成P1", "同上，Priority::P1", "同上", ""),
    ("M", "当日完成P2", "同上，Priority::P2", "同上", ""),
    ("N", "当日完成P3", "同上，Priority::P3", "同上", ""),
    ("O", "当日完成P4", "同上，Priority::P4", "同上", ""),
    ("P", "当日完成合计", "K+L+M+N+O", "计算列", ""),
    ("Q", "当日剩余待验证Bug总量", "当天处于 open 状态（创建日 ≤ 当天 且 未关闭）且最新系统状态为「待修复」的 Bug::dev issue 数", "issue.created_at / closed_at + system note 状态", "同开发看板 O 列"),
    ("R", "累计关闭Bug总量", "截至当天，已关闭的 Bug::dev issue 累计数", "issue.closed_at ≤ 当天", "只增不减"),
    ("S", "高级别(P0+P1)剩余存量", "当天 open 且带 Priority::P0 或 Priority::P1 的 Bug::dev issue 数", "同 Q 列逻辑 + Priority 过滤", "同开发看板 P 列"),
    ("T", "仅前端BUG当日测试通过数", "当天状态变更为「已完成」的 Bug::dev issue 中，按分类规则判定为「仅前端」的数量", "system note + 前后端分类规则", "T+U+V = P"),
    ("U", "仅后端BUG当日测试通过数", "同上，判定为「仅后端」", "同上", ""),
    ("V", "前+后端BUG当日测试通过数", "同上，判定为「前+后端」", "同上", ""),
    ("W", "累计完成功能测试", "截至当天，状态曾变为「已完成」或已关闭的 Function::dev issue 累计数", "遍历 issue 时间线", "只增不减"),
]

for r_data in test_rows:
    for ci, val in enumerate(r_data, 1):
        cell = ws.cell(row, ci, val)
        cell.font = body_font
        cell.alignment = wrap
        cell.border = thin_border
        cell.fill = test_fill
    row += 1

# ===== 列宽 =====
ws.column_dimensions["A"].width = 8
ws.column_dimensions["B"].width = 24
ws.column_dimensions["C"].width = 60
ws.column_dimensions["D"].width = 45
ws.column_dimensions["E"].width = 22

# ===== 保存 =====
try:
    wb.save(EXCEL_PATH)
    print(f"已保存: {EXCEL_PATH}")
except PermissionError:
    alt = EXCEL_PATH.replace("_new.xlsx", "_new3.xlsx")
    wb.save(alt)
    print(f"原文件被占用，已保存到: {alt}")
