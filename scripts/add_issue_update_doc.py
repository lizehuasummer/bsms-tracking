"""
在 Excel 中添加「issue每日更新」说明 sheet，记录前 4 个 sheet 的更新口径。
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

EXCEL_PATH = r"D:\Users\LZH\Documents\opencode\bsms_tracking\【交付HK】BSMS项目整体进度计划 (更新中)20260702备份_new.xlsx"

wb = openpyxl.load_workbook(EXCEL_PATH)

sheet_name = "issue每日更新"
if sheet_name in wb.sheetnames:
    del wb[sheet_name]
ws = wb.create_sheet(sheet_name)

# ===== 样式 =====
title_font = Font(name="微软雅黑", size=14, bold=True, color="FFFFFF")
section_font = Font(name="微软雅黑", size=12, bold=True, color="FFFFFF")
sub_font = Font(name="微软雅黑", size=11, bold=True)
body_font = Font(name="微软雅黑", size=10)
header_font = Font(name="微软雅黑", size=10, bold=True)
title_fill = PatternFill("solid", fgColor="2F5496")
section_fill = PatternFill("solid", fgColor="4472C4")
header_fill = PatternFill("solid", fgColor="D6E4F0")
fill_s1 = PatternFill("solid", fgColor="E2EFDA")
fill_s2 = PatternFill("solid", fgColor="FCE4D6")
fill_s3 = PatternFill("solid", fgColor="D9E2F3")
fill_s4 = PatternFill("solid", fgColor="E4DFEC")
wrap = Alignment(wrap_text=True, vertical="top")
center = Alignment(horizontal="center", vertical="center", wrap_text=True)
thin_border = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)

# ===== Sheet 说明 =====
ROW = 1

def merge_title(text, fill=title_fill):
    global ROW
    ws.merge_cells(start_row=ROW, start_column=1, end_row=ROW, end_column=6)
    cell = ws.cell(ROW, 1, text)
    cell.font = title_font
    cell.fill = fill
    cell.alignment = center
    ws.row_dimensions[ROW].height = 30
    ROW += 2

def section(text, fill=section_fill):
    global ROW
    ws.merge_cells(start_row=ROW, start_column=1, end_row=ROW, end_column=6)
    cell = ws.cell(ROW, 1, text)
    cell.font = section_font
    cell.fill = fill
    cell.alignment = center
    ws.row_dimensions[ROW].height = 26
    ROW += 1

def data_row(cols, fill=None):
    global ROW
    for ci, val in enumerate(cols, 1):
        cell = ws.cell(ROW, ci, val)
        cell.font = body_font
        cell.alignment = wrap
        cell.border = thin_border
        if fill:
            cell.fill = fill
    ws.row_dimensions[ROW].height = max(20, min(80, 15 * len(str(cols[2])) // 40 + 20))
    ROW += 1

def header_row():
    global ROW
    h = ["Sheet", "Sheet名", "行范围", "核心列", "匹配/填充逻辑", "数据来源"]
    for ci, hd in enumerate(h, 1):
        cell = ws.cell(ROW, ci, hd)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = thin_border
    ws.row_dimensions[ROW].height = 22
    ROW += 1

# ===== 标题 =====
merge_title("前 4 个 Sheet 更新要求说明")
ROW += 0

# ===== 通用信息 =====
ws.merge_cells(start_row=ROW, start_column=1, end_row=ROW, end_column=6)
cell = ws.cell(ROW, 1,
    "项目：aladdinx/document/dev-design (ID=4)  |  实例：https://gitlab.stpass.com\n"
    "刷新策略：每次运行脚本时，先通过 --fresh 清除缓存重新抓取，或用缓存加速\n"
    "并发：10 线程并发请求 notes + label_events，每 100 条保存一次缓存")
cell.font = body_font
cell.alignment = wrap
ws.row_dimensions[ROW].height = 40
ROW += 2

# ===== Sheet 1: issue =====
section("Sheet 1: issue —— 全量 Issue 活动明细")
header_row()

issue_desc = [
    ("1", "issue", "第2行起，\n每行1个issue",
     "A:子系统\nB:issue编号\nC:标题\nD:功能点编号\nE~G:来源标签\nH:优先级\nI:最新状态\nJ~L:记录1\nM~O:记录2\n...(最多81组,252列)",
     "获取项目下所有状态的 issue，遍历每个 issue 的\n"
     "notes + label_events，提取 子系统(BSMS::标签)、\n"
     "功能点编号(标题正则提取)、来源标签\n"
     "(Bug::dev / Function::dev / Change::dev / Story::dev)、\n"
     "优先级(Priority::P0~P4)、最新状态(system note\n"
     "最后一条set status to)、所有活动记录\n"
     "(system note正文 / label事件 / 普通评论，\n"
     "按时间排序，每3列一组)。\n"
     "252列 = 9固定列 + 81组×3列",
     "notes.body → J列内容\n"
     "label_event → added/removed label\n"
     "created_at → 时间\n"
     "author.username → 操作人")
]
data_row(issue_desc[0], fill_s1)

ROW += 1

# ===== Sheet 2: Sprint详细安排 =====
section("Sheet 2: Sprint详细安排 —— 功能点开发进度（Function::dev）")
header_row()

sprint_desc = [
    ("2", "Sprint详细安排", "第5行~1166行\n(E列非空为有效行)",
     "E:功能点编号\nF:功能点名称\nI:最新状态\nL:测试中操作人\nM:测试中时间\nN:Process::pass操作人\nO:Process::pass时间\nR:已完成操作人\nS:已完成时间",
     "E列功能点编号 → GitLab issue title 中的\n"
     "XXX-FNNN 匹配。如有多个匹配再按 F列名称匹配。\n\n"
     "I列：system note 最后一条 set status to **XXX**；\n"
     "  若状态=测试中 且当前 labels 含 Process::pass → processpass\n"
     "L/M：system note set status to **测试中** → last author/time\n"
     "N/O：label_event Process::pass + action=add → last user/time\n"
     "R/S：system note set status to **已完成** → last author/time",
     "system notes\n"
     "label_events\n"
     "issue.labels")
]
data_row(sprint_desc[0], fill_s2)

ROW += 1

# ===== Sheet 3: EX开发功能点 =====
section("Sheet 3: EX开发功能点 —— 各子系统功能点（Function::dev）")
header_row()

ex_desc = [
    ("3", "EX开发功能点", "第5行~389行\n(E列非空为有效行)",
     "E:功能点编号\nF:功能点名称\nL:测试中操作人\nM:测试中时间\nN:Process::pass操作人\nO:Process::pass时间\nR:已完成操作人\nS:已完成时间",
     "匹配逻辑同 Sprint详细安排（不填 I 列状态）\n"
     "仅更新 L/M (测试中)、N/O (Process::pass)、\n"
     "R/S (已完成)",
     "同上")
]
data_row(ex_desc[0], fill_s3)

ROW += 1

# ===== Sheet 4: bug =====
section("Sheet 4: bug —— Bug::dev 全量 Bug 数据")
header_row()

bug_desc = [
    ("4", "bug", "先清空第2行起所有数据\n再重新写入全部 Bug::dev",
     "A:子系统\nB:bug编号\nC:bug标题\nD:优先级\nE:最新状态\nF:创建人\nG:创建时间\nH:待修复操作人\nI:待修复时间\nJ:测试中操作人\nK:测试中时间\nL:已完成操作人\nM:已完成时间",
     "获取所有 Bug::dev issue，无需匹配 Excel 现有行\n"
     "直接全部写入。\n\n"
     "A: BSMS::XX → XX\n"
     "B: #iid\n"
     "D: Priority::P0~P4 → P0~P4\n"
     "E: system note 最后一条 set status to 或 issue.state\n"
     "H~M: 对应 status 的 last author/time",
     "system notes\n"
     "issue.labels\n"
     "issue.state")
]
data_row(bug_desc[0], fill_s4)

ROW += 1

# ===== 缓存说明 =====
section("缓存机制")
ROW += 0

cache_info = [
    "issues 缓存 (gitlab_issues_cache.json)：1 小时内有效，超时或 --fresh 时重新抓取",
    "notes 缓存 (gitlab_notes_cache.json)：永久有效，--fresh 时删除重抓；已有缓存跳过",
    "缓存路径：C:\\Users\\EDY\\AppData\\Local\\Temp\\opencode\\",
    "--excel-only：跳过 GitLab 抓取，仅用缓存填表",
    "--fresh：删除两个缓存后重新完整抓取",
]
for ci in cache_info:
    ws.merge_cells(start_row=ROW, start_column=1, end_row=ROW, end_column=6)
    cell = ws.cell(ROW, 1, ci)
    cell.font = body_font
    cell.alignment = wrap
    ROW += 1

ROW += 1

# ===== 前后端分类规则（参考） =====
section("BUG 前/后端分类规则（参考看板列说明）")
rules = [
    "1. Port::fullstack 标签 → 前+后端",
    "2. 本次操作前前后端用户都有活动 → 前+后端",
    "3. 操作人是前端用户：之前有后端活动 → 前+后端；否则 → 仅前端",
    "4. 操作人是后端用户：之前有前端活动 → 前+后端；否则 → 仅后端",
    "5. Port::frontend 且无 Port::backend → 仅前端",
    "6. Port::backend 且无 Port::frontend → 仅后端",
    "7. 以上都不满足 → 前+后端",
]
for rule in rules:
    ws.merge_cells(start_row=ROW, start_column=1, end_row=ROW, end_column=6)
    cell = ws.cell(ROW, 1, rule)
    cell.font = body_font
    ROW += 1

# ===== 列宽 =====
ws.column_dimensions["A"].width = 8
ws.column_dimensions["B"].width = 20
ws.column_dimensions["C"].width = 28
ws.column_dimensions["D"].width = 24
ws.column_dimensions["E"].width = 60
ws.column_dimensions["F"].width = 30

# ===== 保存 =====
try:
    wb.save(EXCEL_PATH)
    print(f"已保存: {EXCEL_PATH}")
except PermissionError:
    alt = EXCEL_PATH.replace(".xlsx", "_backup.xlsx")
    wb.save(alt)
    print(f"已保存到: {alt}")
print("完成！")
